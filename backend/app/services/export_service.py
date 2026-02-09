"""Export service for generating PDF and Excel reports."""

from __future__ import annotations

import io
import logging
from typing import TYPE_CHECKING, Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from app.services.superset_client import SupersetClient

logger = logging.getLogger(__name__)

# Excel style constants
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_BREACH_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_RESOLVED_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")


class ExportService:
    """Generate PDF and Excel exports of dashboard data."""

    def __init__(self, superset_client: SupersetClient):
        self._superset = superset_client

    # ------------------------------------------------------------------
    # PDF generation
    # ------------------------------------------------------------------

    async def generate_pdf(
        self,
        dashboard_id: str,
        filters: dict,
        options: dict,
    ) -> bytes:
        """Generate a PDF report for a dashboard.

        Fetches data for each chart, renders an HTML template,
        and converts to PDF via WeasyPrint.
        """
        from weasyprint import HTML

        chart_data_list = await self._fetch_dashboard_charts(dashboard_id, filters)
        html_content = self._render_pdf_html(dashboard_id, chart_data_list, options)
        pdf_bytes: bytes = HTML(string=html_content).write_pdf()
        return pdf_bytes

    async def _fetch_dashboard_charts(
        self,
        dashboard_id: str,
        filters: dict,
    ) -> list[dict]:
        """Fetch data for all charts in a dashboard."""
        dashboard = await self._superset.get_dashboard(int(dashboard_id))
        chart_ids: list[int] = []
        for chart in dashboard.get("charts", []):
            chart_id = chart.get("id")
            if chart_id is not None:
                chart_ids.append(chart_id)

        results: list[dict] = []
        superset_filters = _dict_to_superset_filters(filters)
        for chart_id in chart_ids:
            data = await self._superset.get_chart_data(chart_id, superset_filters)
            results.append({"chart_id": chart_id, "data": data})
        return results

    def _render_pdf_html(
        self,
        dashboard_id: str,
        chart_data: list[dict],
        options: dict,
    ) -> str:
        """Render an HTML page for PDF conversion."""
        title = options.get("title", f"Dashboard {dashboard_id} Report")
        rows = ""
        for item in chart_data:
            rows += f"<tr><td>Chart {item['chart_id']}</td>"
            query_results = item.get("data", {}).get("result", [])
            row_count = sum(len(qr.get("data", [])) for qr in query_results)
            rows += f"<td>{row_count} rows</td></tr>"

        return f"""<!DOCTYPE html>
<html>
<head><meta charset="utf-8"><title>{title}</title>
<style>
  body {{ font-family: Arial, sans-serif; margin: 40px; }}
  h1 {{ color: #1F4E79; }}
  table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
  th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
  th {{ background: #1F4E79; color: white; }}
</style>
</head>
<body>
  <h1>{title}</h1>
  <table><thead><tr><th>Chart</th><th>Data</th></tr></thead>
  <tbody>{rows}</tbody></table>
</body>
</html>"""

    # ------------------------------------------------------------------
    # Excel generation
    # ------------------------------------------------------------------

    async def generate_excel(
        self,
        dashboard_id: str,
        filters: dict,
        options: dict,
    ) -> bytes:
        """Generate an Excel workbook for a dashboard.

        Sheets: Summary (KPI values), Chart Data (tabular), Detail Data.
        """
        chart_data_list = await self._fetch_dashboard_charts(dashboard_id, filters)
        wb = self._create_excel_workbook(chart_data_list, options)
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def _create_excel_workbook(
        self,
        chart_data: list[dict],
        options: dict,
    ) -> Workbook:
        """Build a formatted Excel workbook."""
        wb = Workbook()

        # --- Sheet 1: Summary ---
        ws_summary = wb.active
        if ws_summary is not None:
            ws_summary.title = "Summary"
            ws_summary.append(["Metric", "Value"])
            _style_header_row(ws_summary, 2)
            ws_summary.append(
                ["Total Charts", len(chart_data)],
            )
            ws_summary.append(
                ["Report Title", options.get("title", "Dashboard Report")],
            )
            ws_summary.column_dimensions["A"].width = 25
            ws_summary.column_dimensions["B"].width = 40

        # --- Sheet 2: Chart Data ---
        ws_charts = wb.create_sheet("Chart Data")
        ws_charts.append(["Chart ID", "Rows", "Status"])
        _style_header_row(ws_charts, 3)
        for item in chart_data:
            query_results = item.get("data", {}).get("result", [])
            row_count = sum(len(qr.get("data", [])) for qr in query_results)
            ws_charts.append([item["chart_id"], row_count, "OK"])
        ws_charts.column_dimensions["A"].width = 15
        ws_charts.column_dimensions["B"].width = 15
        ws_charts.column_dimensions["C"].width = 15

        # --- Sheet 3: Detail Data ---
        ws_detail = wb.create_sheet("Detail Data")
        self._write_detail_sheet(ws_detail, chart_data)

        return wb

    def _write_detail_sheet(
        self,
        ws: Any,
        chart_data: list[dict],
    ) -> None:
        """Write raw data rows to the detail sheet."""
        row_offset = 1
        for item in chart_data:
            query_results = item.get("data", {}).get("result", [])
            for qr in query_results:
                records: list[dict] = qr.get("data", [])
                if not records:
                    continue
                headers = list(records[0].keys())
                # Write header
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_offset, column=col_idx, value=header)
                    cell.font = _HEADER_FONT
                    cell.fill = _HEADER_FILL
                    cell.alignment = _HEADER_ALIGN
                row_offset += 1
                # Write data rows with conditional formatting
                for record in records:
                    for col_idx, header in enumerate(headers, 1):
                        cell = ws.cell(
                            row=row_offset,
                            column=col_idx,
                            value=record.get(header),
                        )
                        status = str(record.get("status", "")).lower()
                        if status in ("breach", "unmatched", "failed"):
                            cell.fill = _BREACH_FILL
                        elif status in ("resolved", "matched", "passed"):
                            cell.fill = _RESOLVED_FILL
                    row_offset += 1
                # Auto-width columns
                for col_idx, header in enumerate(headers, 1):
                    ws.column_dimensions[get_column_letter(col_idx)].width = max(
                        len(str(header)) + 4, 12
                    )
                row_offset += 1  # blank row between charts
        # Freeze top row
        ws.freeze_panes = "A2"


def _style_header_row(ws: Any, num_cols: int) -> None:
    """Apply header styling to the first row of a worksheet."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN


def _dict_to_superset_filters(filters: dict) -> list[dict]:
    """Quick conversion from flat filter dict to Superset filter list."""
    from app.models.filters import GlobalFilters
    from app.services.filter_converter import to_superset_filters

    global_filters = GlobalFilters(**filters) if filters else GlobalFilters()
    return [f.model_dump() for f in to_superset_filters(global_filters)]
